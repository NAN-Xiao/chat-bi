"""
脚本说明：验证项目数据字典 Excel 可以作为埋点和语义配置的维护入口。
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from apps.chat.models.chat_model import AiModelQuestion
from apps.system.crud.tracking_config import (
    build_tracking_prompt_context,
    filter_tracking_config_for_physical_schema,
)
from apps.system.crud.tracking_excel import (
    PhysicalFieldInfo,
    PhysicalTableInfo,
    parse_tracking_excel,
    tracking_config_excel,
)
from apps.system.crud.tracking_expression import compile_tracking_json_expression
from apps.system.schemas.tenant_schema import (
    TenantTrackingConfigDTO,
    TenantTrackingEventGroupDTO,
    TenantTrackingFieldDTO,
    TenantTrackingTableDTO,
)


def _physical_schema() -> dict[str, PhysicalTableInfo]:
    return {
        "event_log": PhysicalTableInfo(
            table_name="event_log",
            table_comment="事件明细表",
            fields=[
                PhysicalFieldInfo("uid", "varchar", "用户ID"),
                PhysicalFieldInfo("event_name", "varchar", "事件名"),
                PhysicalFieldInfo("event_time", "timestamp", "事件时间"),
                PhysicalFieldInfo("event_props", "jsonb", "事件属性"),
            ],
        ),
        "user_profile": PhysicalTableInfo(
            table_name="user_profile",
            table_comment="用户属性表",
            fields=[
                PhysicalFieldInfo("uid", "varchar", "用户ID"),
                PhysicalFieldInfo("country", "varchar", "注册国家"),
            ],
        ),
    }


def _event_physical_schema() -> dict[str, PhysicalTableInfo]:
    return {
        "event": PhysicalTableInfo(
            table_name="event",
            table_comment="事件明细表",
            fields=[
                PhysicalFieldInfo("uid", "varchar", "用户ID"),
                PhysicalFieldInfo("event", "varchar", "事件名"),
                PhysicalFieldInfo("time", "bigint", "事件时间"),
                PhysicalFieldInfo("ext", "json", "事件属性"),
            ],
        ),
    }


def _event_user_physical_schema() -> dict[str, PhysicalTableInfo]:
    return {
        "event": PhysicalTableInfo(
            table_name="event",
            table_comment="事件明细表",
            fields=[
                PhysicalFieldInfo("uid", "varchar", "用户ID"),
                PhysicalFieldInfo("event", "varchar", "事件名"),
                PhysicalFieldInfo("time", "bigint", "事件时间"),
                PhysicalFieldInfo("ext", "json", "事件参数"),
            ],
        ),
        "user": PhysicalTableInfo(
            table_name="user",
            table_comment="用户表",
            fields=[
                PhysicalFieldInfo("account_id", "varchar", "账号ID"),
                PhysicalFieldInfo("hero_detail", "json", "拥有的卡牌明细"),
                PhysicalFieldInfo("total_revenue", "numeric", "累计付费金额"),
            ],
        ),
    }


JSON_FIELD_PARSING_HEADERS = [
    "来源表",
    "来源字段",
    "JSON路径",
    "生成字段名",
    "类型",
    "属性说明",
]
LEGACY_JSON_FIELD_PARSING_HEADERS = JSON_FIELD_PARSING_HEADERS[1:]


def _json_overview_physical_schema() -> dict[str, PhysicalTableInfo]:
    return {
        "event": PhysicalTableInfo(
            table_name="event",
            fields=[
                PhysicalFieldInfo("uid", "varchar", "用户ID"),
                PhysicalFieldInfo("event", "varchar", "事件名"),
                PhysicalFieldInfo("userinfo", "text", "用户信息 JSON"),
                PhysicalFieldInfo("event_only", "text", "事件独有 JSON"),
            ],
        ),
        "user": PhysicalTableInfo(
            table_name="user",
            fields=[
                PhysicalFieldInfo("uid", "varchar", "用户ID"),
                PhysicalFieldInfo("userinfo", "text", "用户信息 JSON"),
                PhysicalFieldInfo("profile_json", "text", "用户独有 JSON"),
            ],
        ),
    }


def _json_overview_workbook(
    rows: list[list[object]],
    *,
    headers: list[str] = JSON_FIELD_PARSING_HEADERS,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "JSON字段解析"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_json_field_parsing_sheet_imports_fields_and_prefers_event_table() -> None:
    parsed = parse_tracking_excel(
        _json_overview_workbook(
            [
                ["", "userinfo", "$._appVersion", "userinfo._appVersion", "文本", "应用版本"],
                ["", "profile_json", "$.country", "profile_json.country", "文本", "注册国家"],
            ]
        ),
        TenantTrackingConfigDTO(
            tenant_id=2001,
            enabled=True,
            default_event_table="event",
        ),
        physical_schema=_json_overview_physical_schema(),
        datasource_type="mysql",
    )

    fields = {(item.table_name, item.field_name): item for item in parsed.editor.fields}
    app_version = fields[("event", "userinfo._appVersion")]
    assert app_version.source_field == "userinfo"
    assert app_version.json_path == "$._appVersion"
    assert app_version.semantic_type == "text"
    assert app_version.field_comment == "应用版本"
    assert ("user", "profile_json.country") in fields
    assert not any(item.table_name == "JSON字段解析" for item in parsed.editor.tables)

    expression = compile_tracking_json_expression(
        app_version.table_name,
        app_version.source_field,
        app_version.json_path,
        app_version.semantic_type,
        "mysql",
    )
    assert expression == "JSON_UNQUOTE(JSON_EXTRACT(`event`.`userinfo`, '$._appVersion'))"


def _json_overview_with_event_field(*, event_type: str, overview_type: str) -> bytes:
    workbook = Workbook()
    event_sheet = workbook.active
    event_sheet.title = "event"
    event_sheet.append(
        [
            "属性名（必填）",
            "属性显示名",
            "属性类型（必填）",
            "更新方式",
            "属性说明",
            "属性标签",
        ]
    )
    event_sheet.append(["userinfo.country", "国家", event_type, "", "国家", "用户"])
    overview = workbook.create_sheet("JSON字段解析")
    overview.append(JSON_FIELD_PARSING_HEADERS)
    overview.append(["", "userinfo", "$.country", "userinfo.country", overview_type, "注册国家"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_json_field_parsing_sheet_skips_incomplete_rows_and_preserves_null_type() -> None:
    parsed = parse_tracking_excel(
        _json_overview_workbook(
            [
                ["", "userinfo", "$._appVersion", "userinfo._appVersion", "空值", "尚无非空样本"],
                ["", "userinfo", "", "", "文本", "空对象，暂不生成字段"],
            ]
        ),
        TenantTrackingConfigDTO(
            tenant_id=2001,
            enabled=True,
            default_event_table="event",
        ),
        physical_schema=_json_overview_physical_schema(),
        datasource_type="mysql",
    )

    field = next(item for item in parsed.editor.fields if item.field_name == "userinfo._appVersion")
    assert field.semantic_type == "text"
    assert field.extra_properties["json_observed_type"] == "空值"
    assert parsed.skipped_rows == 1
    assert any("第 3 行" in warning and "已跳过" in warning for warning in parsed.warnings)


def test_json_field_parsing_sheet_rejects_generated_name_mismatch() -> None:
    with pytest.raises(ValueError, match=r"第 2 行.*生成字段名.*userinfo\.country"):
        parse_tracking_excel(
            _json_overview_workbook(
                [["", "userinfo", "$.country", "userinfo.wrong", "文本", "国家"]]
            ),
            TenantTrackingConfigDTO(
                tenant_id=2001,
                enabled=True,
                default_event_table="event",
            ),
            physical_schema=_json_overview_physical_schema(),
            datasource_type="mysql",
        )


def test_json_field_parsing_sheet_rejects_invalid_json_path() -> None:
    with pytest.raises(ValueError, match=r"第 2 行.*JSON路径.*无效"):
        parse_tracking_excel(
            _json_overview_workbook(
                [["", "userinfo", "$.country[", "userinfo.country", "文本", "国家"]]
            ),
            TenantTrackingConfigDTO(
                tenant_id=2001,
                enabled=True,
                default_event_table="event",
            ),
            physical_schema=_json_overview_physical_schema(),
            datasource_type="mysql",
        )


def test_json_field_parsing_sheet_rejects_ambiguous_source_without_event_candidate() -> None:
    schema = {
        "profile_a": PhysicalTableInfo(
            "profile_a", fields=[PhysicalFieldInfo("payload", "text")]
        ),
        "profile_b": PhysicalTableInfo(
            "profile_b", fields=[PhysicalFieldInfo("payload", "text")]
        ),
    }
    with pytest.raises(ValueError, match=r"第 2 行.*profile_a.*profile_b"):
        parse_tracking_excel(
            _json_overview_workbook(
                [["", "payload", "$.country", "payload.country", "文本", "国家"]]
            ),
            TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
            physical_schema=schema,
            datasource_type="mysql",
        )


def test_json_field_parsing_sheet_honors_explicit_source_table() -> None:
    parsed = parse_tracking_excel(
        _json_overview_workbook(
            [
                ["event", "userinfo", "$.country", "userinfo.country", "文本", "事件国家"],
                ["user", "userinfo", "$.country_code", "userinfo.country_code", "文本", "注册国家码"],
            ]
        ),
        TenantTrackingConfigDTO(
            tenant_id=2001,
            enabled=True,
            default_event_table="event",
        ),
        physical_schema=_json_overview_physical_schema(),
        datasource_type="mysql",
    )

    fields = {(item.table_name, item.field_name): item for item in parsed.editor.fields}
    assert fields[("event", "userinfo.country")].json_path == "$.country"
    assert fields[("user", "userinfo.country_code")].json_path == "$.country_code"


@pytest.mark.parametrize(
    ("table_name", "source_field", "message"),
    [
        ("missing", "userinfo", r"第 2 行.*来源表 missing.*不存在"),
        ("user", "event_only", r"第 2 行.*来源表 user.*不包含来源字段 event_only"),
    ],
)
def test_json_field_parsing_sheet_rejects_invalid_explicit_source_table(
    table_name: str,
    source_field: str,
    message: str,
) -> None:
    with pytest.raises(ValueError, match=message):
        parse_tracking_excel(
            _json_overview_workbook(
                [
                    [
                        table_name,
                        source_field,
                        "$.country",
                        f"{source_field}.country",
                        "文本",
                        "国家",
                    ]
                ]
            ),
            TenantTrackingConfigDTO(
                tenant_id=2001,
                enabled=True,
                default_event_table="event",
            ),
            physical_schema=_json_overview_physical_schema(),
            datasource_type="mysql",
        )


def test_json_field_parsing_sheet_legacy_five_columns_still_prefers_event() -> None:
    parsed = parse_tracking_excel(
        _json_overview_workbook(
            [["userinfo", "$.country", "userinfo.country", "文本", "国家"]],
            headers=LEGACY_JSON_FIELD_PARSING_HEADERS,
        ),
        TenantTrackingConfigDTO(
            tenant_id=2001,
            enabled=True,
            default_event_table="event",
        ),
        physical_schema=_json_overview_physical_schema(),
        datasource_type="mysql",
    )

    field = next(
        item for item in parsed.editor.fields if item.field_name == "userinfo.country"
    )
    assert field.table_name == "event"


def test_json_field_parsing_sheet_deduplicates_matching_table_definition() -> None:
    parsed = parse_tracking_excel(
        _json_overview_with_event_field(event_type="文本", overview_type="文本"),
        TenantTrackingConfigDTO(
            tenant_id=2001,
            enabled=True,
            default_event_table="event",
        ),
        physical_schema=_json_overview_physical_schema(),
        datasource_type="mysql",
    )

    matches = [item for item in parsed.editor.fields if item.field_name == "userinfo.country"]
    assert len(matches) == 1
    assert matches[0].field_comment == "注册国家"


def test_json_field_parsing_sheet_rejects_conflicting_table_definition() -> None:
    with pytest.raises(ValueError, match="定义冲突"):
        parse_tracking_excel(
            _json_overview_with_event_field(event_type="数值", overview_type="文本"),
            TenantTrackingConfigDTO(
                tenant_id=2001,
                enabled=True,
                default_event_table="event",
            ),
            physical_schema=_json_overview_physical_schema(),
            datasource_type="mysql",
        )


def _tracking_config() -> TenantTrackingConfigDTO:
    return TenantTrackingConfigDTO(
        tenant_id=2001,
        enabled=True,
        default_event_table="event_log",
        default_subject_field="uid",
        default_event_name_field="event_name",
        default_event_time_field="event_time",
        field_role_mappings=[{"field_role": "subject_id", "description": "主体用户字段"}],
        sql_rules="支付金额统一使用 event_props.amount；活跃用户按 uid 去重。",
        notes="当前项目以 Excel 维护事件、公共属性、用户属性。",
        tables=[
            TenantTrackingTableDTO(
                tenant_id=2001,
                table_name="event_log",
                table_comment="事件明细表",
                table_role="event_fact",
                aliases=["事件明细"],
            ),
            TenantTrackingTableDTO(
                tenant_id=2001,
                table_name="user_profile",
                table_comment="用户属性表",
                table_role="user_profile",
                aliases=["用户属性"],
            ),
        ],
        fields=[
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="event_log",
                field_name="uid",
                field_comment="用户唯一标识",
                field_role="subject_id",
                semantic_type="identifier",
                aliases=["用户ID"],
            ),
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="event_log",
                field_name="event_name",
                field_comment="业务事件名",
                field_role="event_name",
                semantic_type="text",
                aliases=["事件名"],
            ),
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="event_log",
                field_name="event_props.amount",
                field_comment="支付金额",
                field_role="json_path_metric",
                semantic_type="number",
                source_field="event_props",
                json_path="$.amount",
                aliases=["支付金额"],
                expression="NULLIF((\"event_log\".\"event_props\"::jsonb #>> '{amount}'), '')::numeric",
            ),
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="event_log",
                field_name="event_props.battleResult",
                field_comment="战斗结果",
                field_role="json_path_dimension",
                semantic_type="text",
                source_field="event_props",
                json_path="$.battleResult",
                aliases=["战斗结果"],
                value_mappings=[
                    {"value": "1", "display_name": "胜利", "category": "战斗结果", "description": "战斗胜利", "aliases": ["win"]},
                    {"value": "3", "display_name": "失败", "category": "战斗结果"},
                ],
            ),
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="user_profile",
                field_name="country",
                field_comment="注册国家",
                field_role="profile_dimension",
                semantic_type="text",
                aliases=["注册国家"],
                value_mappings=[
                    {"value": "US", "display_name": "美国", "category": "北美", "description": "美国用户"},
                    {"value": "CN", "display_name": "中国", "category": "东亚"},
                ],
            ),
        ],
        event_name_mappings=[
            {
                "event_name": "login",
                "event_display_name": "登录",
                "event_category": "基础事件",
                "description": "用户登录事件",
                "aliases": ["登录事件"],
                "properties": [
                    {
                        "property_name": "duration",
                        "property_display_name": "停留时长",
                        "property_type": "数值",
                        "source_field": "event_props",
                        "json_path": "$.duration",
                        "description": "本次停留秒数",
                    }
                ],
            },
            {
                "metric": "付费相关事件",
                "events": ["pay_success", "refund_success"],
                "description": "付费与退款事件",
            },
        ],
    )


def _headers(workbook_bytes: bytes, sheet_name: str) -> list[str]:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    return [str(cell or "") for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]


def _sheet_rows(workbook_bytes: bytes, sheet_name: str) -> list[tuple[object, ...]]:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    return list(sheet.iter_rows(min_row=2, values_only=True))


EVENT_GROUP_HEADERS = [
    "分组标识（必填）",
    "分组名称（必填）",
    "分组说明",
    "分组排序",
    "事件名（必填）",
    "事件排序",
    "启用",
]


def _workbook_with_event_group_rows(rows: list[list[object]]) -> bytes:
    workbook_bytes = tracking_config_excel(_tracking_config(), physical_schema=_physical_schema()).getvalue()
    workbook = load_workbook(BytesIO(workbook_bytes))
    sheet = workbook["事件分组"]
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_event_groups_export_as_independent_sheet() -> None:
    config = _tracking_config().model_copy(
        update={
            "event_groups": [
                TenantTrackingEventGroupDTO(
                    tenant_id=2001,
                    group_key="payment_process",
                    group_name="支付流程事件",
                    description="只统计流程量",
                    event_names=["pay_success", "refund_success"],
                    sort_order=10,
                )
            ]
        }
    )

    workbook_bytes = tracking_config_excel(config, physical_schema=_physical_schema()).getvalue()

    assert _headers(workbook_bytes, "事件分组") == EVENT_GROUP_HEADERS
    assert _sheet_rows(workbook_bytes, "事件分组") == [
        ("payment_process", "支付流程事件", "只统计流程量", 10, "pay_success", 1, "是"),
        ("payment_process", "支付流程事件", "只统计流程量", 10, "refund_success", 2, "是"),
    ]
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=False, data_only=True)
    sheet = workbook["事件分组"]
    assert sheet.column_dimensions["A"].width >= 18
    assert sheet.column_dimensions["A"].max >= 2


def test_event_parameter_mapping_exports_camel_case_event_name() -> None:
    config = _tracking_config().model_copy(
        update={
            "event_name_mappings": [
                {
                    "eventName": "pay_success",
                    "event_display_name": "支付成功",
                    "properties": [],
                }
            ]
        }
    )

    workbook_bytes = tracking_config_excel(config, physical_schema=_physical_schema()).getvalue()

    assert _sheet_rows(workbook_bytes, "事件参数对照")[0][0] == "pay_success"


def test_event_group_sheet_imports_authoritative_groups() -> None:
    content = _workbook_with_event_group_rows(
        [
            ["payment_process", "支付流程事件", "只统计流程量", 10, "refund_success", 20, "是"],
            ["payment_process", "支付流程事件", "只统计流程量", 10, "pay_success", 10, "是"],
        ]
    )

    parsed = parse_tracking_excel(content, _tracking_config(), physical_schema=_physical_schema())

    assert parsed.editor.event_groups is not None
    assert len(parsed.editor.event_groups) == 1
    group = parsed.editor.event_groups[0]
    assert group.group_key == "payment_process"
    assert group.event_names == ["pay_success", "refund_success"]
    assert group.sort_order == 10


def test_empty_event_group_sheet_preserves_existing_groups() -> None:
    content = _workbook_with_event_group_rows([])

    parsed = parse_tracking_excel(content, _tracking_config(), physical_schema=_physical_schema())

    assert parsed.editor.event_groups is None


def test_event_group_sheet_rejects_unknown_event() -> None:
    content = _workbook_with_event_group_rows(
        [["payment_process", "支付流程事件", "只统计流程量", 10, "PayFinish2", 10, "是"]]
    )

    with pytest.raises(ValueError, match="第 2 行.*PayFinish2"):
        parse_tracking_excel(content, _tracking_config(), physical_schema=_physical_schema())


def test_event_group_sheet_rejects_unrecognized_header_with_data() -> None:
    content = _workbook_with_event_group_rows(
        [["payment_process", "支付流程事件", "只统计流程量", 10, "pay_success", 10, "是"]]
    )
    workbook = load_workbook(BytesIO(content))
    sheet = workbook["事件分组"]
    for column_index in range(1, len(EVENT_GROUP_HEADERS) + 1):
        sheet.cell(1, column_index).value = f"错误列{column_index}"
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(ValueError, match="事件分组.*表头无法识别"):
        parse_tracking_excel(output.getvalue(), _tracking_config(), physical_schema=_physical_schema())


def test_event_group_sheet_with_unrecognized_header_but_no_data_is_ignored() -> None:
    content = _workbook_with_event_group_rows([])
    workbook = load_workbook(BytesIO(content))
    sheet = workbook["事件分组"]
    for column_index in range(1, len(EVENT_GROUP_HEADERS) + 1):
        sheet.cell(1, column_index).value = f"错误列{column_index}"
    output = BytesIO()
    workbook.save(output)

    parsed = parse_tracking_excel(
        output.getvalue(),
        _tracking_config(),
        physical_schema=_physical_schema(),
    )

    assert parsed.editor.event_groups is None


def test_event_group_sheet_reports_row_for_invalid_group_key() -> None:
    content = _workbook_with_event_group_rows(
        [["Payment-Process", "支付流程事件", "只统计流程量", 10, "pay_success", 10, "是"]]
    )

    with pytest.raises(ValueError, match="第 2 行.*分组标识"):
        parse_tracking_excel(content, _tracking_config(), physical_schema=_physical_schema())


def test_event_dictionary_import_rejects_removing_group_member() -> None:
    existing = _tracking_config().model_copy(
        update={
            "event_groups": [
                TenantTrackingEventGroupDTO(
                    tenant_id=2001,
                    group_key="payment_process",
                    group_name="支付流程事件",
                    event_names=["pay_success"],
                )
            ]
        }
    )
    workbook_bytes = tracking_config_excel(existing, physical_schema=_physical_schema()).getvalue()
    workbook = load_workbook(BytesIO(workbook_bytes))
    workbook.remove(workbook["事件分组"])
    event_parameter_sheet = workbook["事件参数对照"]
    for row_index in range(event_parameter_sheet.max_row, 1, -1):
        if event_parameter_sheet.cell(row_index, 1).value == "pay_success":
            event_parameter_sheet.delete_rows(row_index, 1)
    event_sheet = workbook["event_log"]
    for row_index in range(event_sheet.max_row, 1, -1):
        if event_sheet.cell(row_index, 11).value == "pay_success":
            event_sheet.delete_rows(row_index, 1)
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(ValueError, match="payment_process.*pay_success"):
        parse_tracking_excel(output.getvalue(), existing, physical_schema=_physical_schema())


def test_excel_unknown_columns_roundtrip_as_extra_properties() -> None:
    """
    是什么：验证数据字典 Excel 中未被系统识别的扩展列，导入后会保留并在导出时还原。
    """
    workbook_bytes = tracking_config_excel(_tracking_config(), physical_schema=_physical_schema()).getvalue()
    workbook = load_workbook(BytesIO(workbook_bytes))

    table_sheet = workbook["_表映射"]
    table_extra_col = table_sheet.max_column + 1
    table_sheet.cell(1, table_extra_col).value = "扩展_负责人"
    for row_index in range(2, table_sheet.max_row + 1):
        if table_sheet.cell(row_index, 2).value == "event_log":
            table_sheet.cell(row_index, table_extra_col).value = "数仓团队"
            break

    field_sheet = workbook["event_log"]
    field_extra_col = field_sheet.max_column + 1
    field_sheet.cell(1, field_extra_col).value = "扩展_业务线"
    for row_index in range(2, field_sheet.max_row + 1):
        if field_sheet.cell(row_index, 2).value == "event_props.amount":
            field_sheet.cell(row_index, field_extra_col).value = "商业化"
            break

    output = BytesIO()
    workbook.save(output)

    parsed = parse_tracking_excel(
        output.getvalue(),
        TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
        physical_schema=_physical_schema(),
        datasource_type="postgresql",
    )

    table = next(item for item in parsed.editor.tables if item.table_name == "event_log")
    field = next(item for item in parsed.editor.fields if item.table_name == "event_log" and item.field_name == "event_props.amount")
    assert table.extra_properties == {"扩展_负责人": "数仓团队"}
    assert field.extra_properties == {"扩展_业务线": "商业化"}

    exported = tracking_config_excel(parsed.editor, physical_schema=_physical_schema()).getvalue()
    table_headers = _headers(exported, "_表映射")
    field_headers = _headers(exported, "event_log")
    assert table_headers[-1] == "扩展_负责人"
    assert field_headers[-1] == "扩展_业务线"

    table_owner_index = table_headers.index("扩展_负责人")
    field_domain_index = field_headers.index("扩展_业务线")
    assert any(row[1] == "event_log" and row[table_owner_index] == "数仓团队" for row in _sheet_rows(exported, "_表映射"))
    assert any(row[1] == "event_props.amount" and row[field_domain_index] == "商业化" for row in _sheet_rows(exported, "event_log"))


LEGACY_BUSINESS_HEADERS = [
    "row_type",
    "field_name",
    "field_display_name",
    "field_type",
    "field_role",
    "semantic_type",
    "source_field",
    "json_path",
    "expression",
    "required",
    "value",
    "value_display_name",
    "value_category",
    "collect_side",
    "event_name",
    "event_display_name",
    "event_category",
    "example_values",
    "description",
    "ai_notes",
    "aliases",
]


def _legacy_business_workbook(sheet_name: str, rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(LEGACY_BUSINESS_HEADERS)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_event_and_user_tables_export_attribute_sheet_format() -> None:
    """
    是什么：验证真实 event/user 表导出为业务属性表格式，不再暴露 row_type 等内部列。
    """
    config = TenantTrackingConfigDTO(
        tenant_id=2001,
        enabled=True,
        default_event_table="event",
        default_subject_field="uid",
        default_event_name_field="event",
        default_event_time_field="time",
        tables=[
            TenantTrackingTableDTO(tenant_id=2001, table_name="event", table_role="event_fact", aliases=["事件表"]),
            TenantTrackingTableDTO(tenant_id=2001, table_name="user", table_role="user_profile", aliases=["用户表"]),
        ],
        fields=[
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="event",
                field_name="ext.battleResult",
                field_comment="战斗结果字段",
                field_role="json_path_dimension",
                semantic_type="text",
                source_field="ext",
                json_path="$.battleResult",
                aliases=["战斗结果"],
            ),
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="user",
                field_name="total_revenue",
                field_comment="累计付费金额，每次充值成功时累加，单位：分",
                semantic_type="number",
                aliases=["累计付费金额"],
                update_mode="user_add",
            ),
        ],
    )

    workbook_bytes = tracking_config_excel(config, physical_schema=_event_user_physical_schema()).getvalue()

    expected_headers = ["属性名（必填）", "属性显示名", "属性类型（必填）", "更新方式", "属性说明", "属性标签"]
    assert _headers(workbook_bytes, "event") == expected_headers
    assert _headers(workbook_bytes, "user") == expected_headers
    event_rows = _sheet_rows(workbook_bytes, "event")
    user_rows = _sheet_rows(workbook_bytes, "user")
    json_rows = _sheet_rows(workbook_bytes, "JSON字段解析")
    assert not any(row[0] == "ext.battleResult" for row in event_rows)
    assert any(
        row[0] == "ext"
        and row[1] == "$.battleResult"
        and row[2] == "ext.battleResult"
        and row[3] == "文本"
        for row in json_rows
    )
    assert any(row[0] == "total_revenue" and row[1] == "累计付费金额" and row[2] == "数值" and row[3] == "user_add" for row in user_rows)
    assert "row_type" not in _headers(workbook_bytes, "event")
    assert "row_type" not in _headers(workbook_bytes, "user")


def test_event_table_exports_event_parameter_mapping_sheet() -> None:
    """
    是什么：验证 event 表属性导出时，事件参数会单独放入“事件参数对照”sheet。
    """
    config = TenantTrackingConfigDTO(
        tenant_id=2001,
        enabled=True,
        default_event_table="event",
        default_subject_field="uid",
        default_event_name_field="event",
        default_event_time_field="time",
        tables=[
            TenantTrackingTableDTO(tenant_id=2001, table_name="event", table_role="event_fact", aliases=["事件表"]),
            TenantTrackingTableDTO(tenant_id=2001, table_name="user", table_role="user_profile", aliases=["用户表"]),
        ],
        fields=[
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="event",
                field_name="ext.battleResult",
                field_comment="战斗结果，固定值win",
                field_role="json_path_dimension",
                semantic_type="text",
                source_field="ext",
                json_path="$.battleResult",
                aliases=["战斗结果"],
                category="战斗",
            ),
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="event",
                field_name="ext.hero_info.hero_level",
                field_comment="卡牌等级",
                field_role="json_path_dimension",
                semantic_type="number",
                source_field="ext",
                json_path="$.hero_info.hero_level",
                aliases=["卡牌等级"],
                category="战斗",
            ),
        ],
        event_name_mappings=[
            {
                "event_name": "BattleEnd",
                "event_display_name": "战斗结束",
                "event_category": "战斗",
                "properties": [
                    {
                        "property_name": "battleResult",
                        "property_display_name": "战斗结果",
                        "property_type": "文本",
                        "source_field": "ext",
                        "json_path": "$.battleResult",
                        "description": "战斗结果，固定值win",
                    },
                    {
                        "property_name": "hero_info.hero_level",
                        "property_display_name": "卡牌等级",
                        "property_type": "数值",
                        "source_field": "ext",
                        "json_path": "$.hero_info.hero_level",
                        "description": "卡牌等级",
                    },
                ],
            }
        ],
    )

    workbook_bytes = tracking_config_excel(config, physical_schema=_event_user_physical_schema()).getvalue()

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    assert "事件参数对照" in workbook.sheetnames
    assert _headers(workbook_bytes, "事件参数对照") == [
        "事件名（必填）",
        "事件显示名",
        "事件说明",
        "事件标签",
        "数据源字段",
        "属性名（必填）",
        "属性显示名",
        "属性类型（必填）",
        "属性说明",
    ]
    rows = _sheet_rows(workbook_bytes, "事件参数对照")
    assert any(
        row[0] == "BattleEnd"
        and row[1] == "战斗结束"
        and row[3] == "战斗"
        and row[4] == "ext"
        and row[5] == "battleResult"
        and row[6] == "战斗结果"
        and row[7] == "文本"
        and row[8] == "战斗结果，固定值win"
        for row in rows
    )
    assert any(row[5] == "hero_info.hero_level" and row[6] == "卡牌等级" and row[7] == "数值" for row in rows)
    styled_workbook = load_workbook(BytesIO(workbook_bytes), read_only=False, data_only=True)
    styled_sheet = styled_workbook["事件参数对照"]
    merged_ranges = {str(merged_range) for merged_range in styled_sheet.merged_cells.ranges}
    assert {"A2:A3", "B2:B3", "C2:C3", "D2:D3"}.issubset(merged_ranges)
    assert styled_sheet["A2"].value == "BattleEnd"
    assert styled_sheet["A3"].value is None
    assert styled_sheet["E2"].value == "ext"
    assert styled_sheet["E3"].value == "ext"


def test_event_parameter_mapping_sheet_imports_event_properties() -> None:
    """
    是什么：验证“事件参数对照”sheet 可以导入为事件与参数关系。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "事件参数对照"
    sheet.append([
        "事件名（必填）",
        "事件显示名",
        "事件说明",
        "事件标签",
        "采集端",
        "数据源字段",
        "属性名（必填）",
        "属性显示名",
        "属性类型（必填）",
        "属性说明",
    ])
    sheet.append(["BattleEnd", "战斗结束", "战斗结算时上报", "战斗", "server", "ext", "battleResult", "战斗结果", "文本", "战斗结果字段"])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_tracking_excel(
        output.getvalue(),
        TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
        physical_schema=_event_user_physical_schema(),
        datasource_type="mysql",
    )

    event = next(item for item in parsed.editor.event_name_mappings if isinstance(item, dict) and item.get("event_name") == "BattleEnd")
    assert event["event_display_name"] == "战斗结束"
    assert event["event_category"] == "战斗"
    assert "collect_side" not in event
    prop = event["properties"][0]
    assert prop["property_name"] == "ext.battleResult"
    assert prop["property_display_name"] == "战斗结果"
    assert prop["property_type"] == "文本"
    assert prop["source_field"] == "ext"
    assert prop["json_path"] == "$.battleResult"


def test_event_parameter_mapping_compact_rows_inherit_event_context() -> None:
    """
    是什么：验证“事件参数对照”中后续属性行省略事件名时，会继承上一条事件上下文。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "事件参数对照"
    sheet.append([
        "事件名（必填）",
        "事件显示名",
        "事件说明",
        "事件标签",
        "采集端",
        "数据源字段",
        "属性名（必填）",
        "属性显示名",
        "属性类型（必填）",
        "属性说明",
    ])
    sheet.append(["achievement_finish", "达成成就", "达成成就后上报", "任务与活动", "client", "ext", "achievement_id", "成就ID", "文本", "成就ID，建议挂维度表"])
    sheet.append(["", "", "", "", "", "", "achievement_type", "成就类型", "文本", "成就类型枚举：collection(收集)/battle(战斗)/level(等级)/social(社交)等"])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_tracking_excel(
        output.getvalue(),
        TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
        physical_schema=_event_user_physical_schema(),
        datasource_type="mysql",
    )

    event = next(item for item in parsed.editor.event_name_mappings if isinstance(item, dict) and item.get("event_name") == "achievement_finish")
    props = event["properties"]
    assert [prop["property_name"] for prop in props] == ["ext.achievement_id", "ext.achievement_type"]
    assert props[1]["property_display_name"] == "成就类型"
    assert props[1]["json_path"] == "$.achievement_type"
    assert not parsed.warnings


def test_event_parameter_mapping_compact_rows_inherit_source_field() -> None:
    """
    是什么：验证“事件参数对照”中同一事件的紧凑属性行，会继承上一行的数据源字段。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "事件参数对照"
    sheet.append([
        "事件名（必填）",
        "事件显示名",
        "事件说明",
        "事件标签",
        "采集端",
        "数据源字段",
        "属性名（必填）",
        "属性显示名",
        "属性类型（必填）",
        "属性说明",
    ])
    sheet.append(["ServerPayLog", "后端充值", "", "Pay", "", "personal", "orderId", "订单ID", "文本", "订单ID"])
    sheet.append(["", "", "", "", "", "", "money", "金额", "数值", "金额"])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_tracking_excel(
        output.getvalue(),
        TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
        physical_schema=_event_user_physical_schema(),
        datasource_type="mysql",
    )

    event = next(item for item in parsed.editor.event_name_mappings if isinstance(item, dict) and item.get("event_name") == "ServerPayLog")
    props = event["properties"]
    assert [prop["property_name"] for prop in props] == ["personal.orderId", "personal.money"]
    assert [prop["source_field"] for prop in props] == ["personal", "personal"]
    assert [prop["json_path"] for prop in props] == ["$.orderId", "$.money"]
    assert not parsed.warnings


def test_event_parameter_mapping_requires_source_field() -> None:
    """来源字段无法推断时必须拒绝导入，不能静默使用 ext。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "事件参数对照"
    sheet.append([
        "事件名（必填）",
        "事件显示名",
        "事件说明",
        "事件标签",
        "数据源字段",
        "属性名（必填）",
        "属性显示名",
        "属性类型（必填）",
        "属性说明",
    ])
    sheet.append(["ServerPayLog", "后端充值", "", "Pay", "", "money", "金额", "数值", "金额"])
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(ValueError) as exc_info:
        parse_tracking_excel(
            output.getvalue(),
            TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
            physical_schema=_event_user_physical_schema(),
            datasource_type="mysql",
        )

    message = str(exc_info.value)
    assert "事件参数对照" in message
    assert "第 2 行" in message
    assert "数据源字段" in message


def test_event_parameter_mapping_does_not_inherit_source_across_events() -> None:
    """切换事件后必须重新指定来源字段。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "事件参数对照"
    sheet.append([
        "事件名（必填）",
        "事件显示名",
        "事件说明",
        "事件标签",
        "数据源字段",
        "属性名（必填）",
        "属性显示名",
        "属性类型（必填）",
        "属性说明",
    ])
    sheet.append(["ServerPayLog", "后端充值", "", "Pay", "personal", "money", "金额", "数值", "金额"])
    sheet.append(["UserActive", "当日活跃", "", "Login", "", "sessionId", "会话ID", "文本", "会话ID"])
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(ValueError, match="第 3 行.*数据源字段"):
        parse_tracking_excel(
            output.getvalue(),
            TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
            physical_schema=_event_user_physical_schema(),
            datasource_type="mysql",
        )


def test_user_attribute_sheet_imports_update_mode_and_json_subfield() -> None:
    """
    是什么：验证 user 表新版属性表格式可以导入更新方式，并能从点号属性名推断 JSON 子字段。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "user"
    sheet.append(["属性名（必填）", "属性显示名", "属性类型（必填）", "更新方式", "属性说明", "JSON路径", "样例", "属性标签"])
    sheet.append(["total_revenue", "累计付费金额", "数值", "user_add", "累计付费金额，每次充值成功时累加，单位：分", "", "", "付费"])
    sheet.append(["hero_detail.hero_level", "卡牌等级", "数值", "user_set", "卡牌等级", "$.hero_level", "12,15", "卡牌"])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_tracking_excel(
        output.getvalue(),
        TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
        physical_schema=_event_user_physical_schema(),
        datasource_type="mysql",
    )

    total_revenue = next(field for field in parsed.editor.fields if field.table_name == "user" and field.field_name == "total_revenue")
    assert total_revenue.semantic_type == "number"
    assert total_revenue.aliases == ["累计付费金额"]
    assert getattr(total_revenue, "update_mode") == "user_add"
    assert getattr(total_revenue, "category") == "付费"
    hero_level = next(field for field in parsed.editor.fields if field.table_name == "user" and field.field_name == "hero_detail.hero_level")
    assert hero_level.source_field == "hero_detail"
    assert hero_level.json_path == "$.hero_level"
    assert getattr(hero_level, "update_mode") == "user_set"
    assert hero_level.example_values == ["12", "15"]


def test_tracking_excel_uses_physical_table_sheets_and_roundtrips_to_prompt_context() -> None:
    """
    是什么：验证导出的 Excel 以物理表 sheet 作为唯一主维护入口，且能导入回项目语义。
    谁调用：跑测试时 pytest 会找到并执行它。
    做了什么：导出、检查 sheet/表头/行类型，再导入并生成 agent prompt context。
    """
    config = _tracking_config()
    workbook_bytes = tracking_config_excel(config, physical_schema=_physical_schema()).getvalue()

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    assert workbook.sheetnames == [
        "_说明",
        "_表映射",
        "事件参数对照",
        "事件分组",
        "JSON字段解析",
        "event_log",
        "user_profile",
        "_SQL规则",
    ]
    assert _headers(workbook_bytes, "事件参数对照") == [
        "事件名（必填）",
        "事件显示名",
        "事件说明",
        "事件标签",
        "数据源字段",
        "属性名（必填）",
        "属性显示名",
        "属性类型（必填）",
        "属性说明",
    ]
    assert _headers(workbook_bytes, "event_log")[:13] == [
        "行类型",
        "字段名",
        "字段显示名",
        "字段类型",
        "字段角色",
        "语义类型",
        "来源字段",
        "JSON 路径",
        "字段表达式",
        "是否必填",
        "字段/事件取值",
        "取值显示名",
        "取值标签",
    ]

    physical_event_rows = _sheet_rows(workbook_bytes, "event_log")
    event_field_row = next(row for row in physical_event_rows if row[1] == "event_name" and row[0] == "physical_field")
    assert event_field_row[3] == "varchar"
    assert event_field_row[4] == "event_name"
    assert event_field_row[5] == "text"
    assert any(row[0] == "event_value" and row[1] == "event_name" and row[10] == "login" for row in physical_event_rows)
    assert any(row[0] == "event_value" and row[1] == "event_name" and row[10] == "pay_success" for row in physical_event_rows)
    assert any(row[0] == "dictionary_field" and row[1] == "event_props.amount" and row[6] == "event_props" and row[7] == "$.amount" for row in physical_event_rows)
    amount_row = next(row for row in physical_event_rows if row[0] == "dictionary_field" and row[1] == "event_props.amount")
    assert amount_row[8] in (None, "")
    battle_row_index = next(
        index
        for index, row in enumerate(physical_event_rows)
        if row[0] == "dictionary_field" and row[1] == "event_props.battleResult" and row[7] == "$.battleResult"
    )
    assert physical_event_rows[battle_row_index + 1][0] == "field_value"
    assert physical_event_rows[battle_row_index + 1][1] == "event_props.battleResult"
    assert physical_event_rows[battle_row_index + 1][10] == "1"
    assert physical_event_rows[battle_row_index + 2][0] == "field_value"
    assert physical_event_rows[battle_row_index + 2][1] == "event_props.battleResult"
    assert physical_event_rows[battle_row_index + 2][10] == "3"

    user_rows = _sheet_rows(workbook_bytes, "user_profile")
    assert any(row[0] == "field_value" and row[1] == "country" and row[10] == "US" and row[11] == "美国" for row in user_rows)

    parsed = parse_tracking_excel(
        workbook_bytes,
        TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
        physical_schema=_physical_schema(),
        datasource_type="postgresql",
    )
    assert parsed.profile == "shuzhi_generic_v1"
    assert not parsed.warnings
    assert parsed.editor.default_event_table == "event_log"
    assert parsed.editor.default_subject_field == "uid"
    assert parsed.editor.notes == "当前项目以 Excel 维护事件、公共属性、用户属性。"
    assert {"field_role": "subject_id", "description": "主体用户字段"} in parsed.editor.field_role_mappings
    assert any(table.table_name == "event_log" and "事件明细" in table.aliases for table in parsed.editor.tables)
    assert any(field.table_name == "event_log" and field.field_name == "event_props.amount" for field in parsed.editor.fields)
    assert any(field.table_name == "event_log" and field.field_name == "event_props.duration" for field in parsed.editor.fields)
    amount_field = next(field for field in parsed.editor.fields if field.table_name == "event_log" and field.field_name == "event_props.amount")
    assert amount_field.expression is None
    assert any(event.get("event_name") == "pay_success" for event in parsed.editor.event_name_mappings if isinstance(event, dict))
    login_event = next(event for event in parsed.editor.event_name_mappings if isinstance(event, dict) and event.get("event_name") == "login")
    assert "登录事件" in login_event.get("aliases", [])
    assert any(prop.get("source_field") == "event_props" and prop.get("json_path") == "$.duration" for prop in login_event.get("properties", []))
    battle_result = next(field for field in parsed.editor.fields if field.table_name == "event_log" and field.field_name == "event_props.battleResult")
    assert {"value": "1", "display_name": "胜利", "category": "战斗结果", "description": "战斗胜利", "aliases": ["win"]} in battle_result.value_mappings
    assert {"value": "3", "display_name": "失败", "category": "战斗结果"} in battle_result.value_mappings
    country = next(field for field in parsed.editor.fields if field.table_name == "user_profile" and field.field_name == "country")
    assert "注册国家" in country.aliases
    assert {"value": "US", "display_name": "美国", "category": "北美", "description": "美国用户"} in country.value_mappings

    context, summary = build_tracking_prompt_context(parsed.editor, datasource_type="postgresql")
    assert "<Workspace-Tracking-Rules>" in context
    assert "event_log.event_props.amount" in context
    assert 'expression=NULLIF(("event_log"."event_props"::jsonb ->> \'amount\'), \'\')::numeric' in context
    assert "pay_success" in context
    assert "支付金额统一使用 event_props.amount" in context
    assert any("event_props.amount" in item for item in summary)


def test_legacy_collect_side_column_is_silently_ignored() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "事件参数对照"
    sheet.append([
        "事件名（必填）",
        "事件显示名",
        "事件说明",
        "事件标签",
        "采集端",
        "数据源字段",
        "属性名（必填）",
        "属性显示名",
        "属性类型（必填）",
        "属性说明",
    ])
    sheet.append([
        "login",
        "用户登录",
        "登录成功事件",
        "基础事件",
        "client",
        "ext",
        "serverId",
        "服务器ID",
        "文本",
        "服务器ID",
    ])
    output = BytesIO()
    workbook.save(output)

    parsed = parse_tracking_excel(
        output.getvalue(),
        TenantTrackingConfigDTO(tenant_id=2001, datasource_id=3, enabled=True),
    )

    event = parsed.editor.event_name_mappings[0]
    assert event["event_name"] == "login"
    assert "collect_side" not in event
    assert "collectSide" not in event


def test_event_sheet_event_value_rows_can_roundtrip_event_dictionary() -> None:
    """
    是什么：验证只维护物理表 sheet 里的 event_value 行，也能导入为事件值字典。
    谁调用：跑测试时 pytest 会找到并执行它。
    做了什么：删除兼容事件 sheet 和枚举 sheet，确保 event_log sheet 仍可回填事件定义。
    """
    config = _tracking_config()
    workbook_bytes = tracking_config_excel(config, physical_schema=_physical_schema()).getvalue()
    workbook = load_workbook(BytesIO(workbook_bytes))
    for sheet_name in list(workbook.sheetnames):
        if sheet_name not in {"_表映射", "event_log"}:
            del workbook[sheet_name]
    output = BytesIO()
    workbook.save(output)

    parsed = parse_tracking_excel(
        output.getvalue(),
        TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
        physical_schema=_physical_schema(),
        datasource_type="postgresql",
    )

    assert any(event.get("event_name") == "login" for event in parsed.editor.event_name_mappings if isinstance(event, dict))
    assert any(event.get("event_name") == "pay_success" for event in parsed.editor.event_name_mappings if isinstance(event, dict))


def test_compact_event_value_rows_inherit_previous_context() -> None:
    """
    是什么：验证 Excel 里后续 event_value 行省略上下文字段时，也不会被导入跳过。
    谁调用：跑测试时 pytest 会找到并执行它。
    做了什么：模拟用户粘贴的 compact 形态，只有第一行写 row_type/field_name。
    """
    config = _tracking_config()
    workbook_bytes = tracking_config_excel(config, physical_schema=_physical_schema()).getvalue()
    workbook = load_workbook(BytesIO(workbook_bytes))
    sheet = workbook["event_log"]
    for row_index in range(2, sheet.max_row + 1):
        if sheet.cell(row_index, 1).value == "event_value" and sheet.cell(row_index, 11).value != "login":
            for col_index in (1, 2, 4, 5, 6):
                sheet.cell(row_index, col_index).value = None
    output = BytesIO()
    workbook.save(output)

    parsed = parse_tracking_excel(
        output.getvalue(),
        TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
        physical_schema=_physical_schema(),
        datasource_type="postgresql",
    )

    assert any(event.get("event_name") == "login" for event in parsed.editor.event_name_mappings if isinstance(event, dict))
    assert any(event.get("event_name") == "pay_success" for event in parsed.editor.event_name_mappings if isinstance(event, dict))


def test_event_table_legacy_row_type_sheet_is_not_imported() -> None:
    """
    是什么：验证 event 表不再兼容旧 row_type 格式导入。
    """
    parsed = parse_tracking_excel(
        _legacy_business_workbook(
            "event",
            [
                ["physical_field", "ext", "事件参数", "json", "", "json"],
                ["dictionary_field", "ext.battleResult", "战斗结果", "text", "json_path_dimension", "text", "ext", "$.battleResult"],
                ["field_value", "", "", "", "", "", "", "", "", "", "1", "胜利", "战斗结果", "", "", "", "", "", "胜利时上报", "", "win"],
                ["", "", "", "", "", "", "", "", "", "", "3", "失败", "战斗结果", "", "", "", "", "", "失败时上报", "", "lose"],
                ["", "", "", "", "", "", "", "", "", "", "4", "平局", "战斗结果"],
            ],
        ),
        TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
        physical_schema=_event_physical_schema(),
        datasource_type="mysql",
    )

    assert not any(field.table_name == "event" for field in parsed.editor.fields)
    assert any("event sheet 使用旧 row_type 格式" in warning for warning in parsed.warnings)


def test_single_event_sheet_can_maintain_event_values_and_defaults() -> None:
    """
    是什么：验证真实 event 表只保留 event sheet 时，也能维护事件值字典。
    谁调用：跑测试时 pytest 会找到并执行它。
    做了什么：删除所有辅助 sheet，只导入 event sheet，检查默认事件表/事件名字段和事件值。
    """
    base = _tracking_config()
    config = base.model_copy(
        update={
            "default_event_table": "event",
            "default_event_name_field": "event",
            "default_event_time_field": "time",
            "tables": [
                TenantTrackingTableDTO(
                    tenant_id=2001,
                    table_name="event",
                    table_comment="事件明细表",
                    table_role="event_fact",
                )
            ],
            "fields": [
                TenantTrackingFieldDTO(
                    tenant_id=2001,
                    table_name="event",
                    field_name="uid",
                    field_comment="用户唯一标识",
                    field_role="subject_id",
                    semantic_type="identifier",
                ),
                TenantTrackingFieldDTO(
                    tenant_id=2001,
                    table_name="event",
                    field_name="event",
                    field_comment="业务事件名",
                    field_role="event_name",
                    semantic_type="text",
                ),
                TenantTrackingFieldDTO(
                    tenant_id=2001,
                    table_name="event",
                    field_name="ext.duration",
                    field_comment="停留时长",
                    field_role="json_path_metric",
                    semantic_type="number",
                    source_field="ext",
                    json_path="$.duration",
                ),
            ],
        }
    )
    workbook_bytes = tracking_config_excel(config, physical_schema=_event_physical_schema()).getvalue()
    workbook = load_workbook(BytesIO(workbook_bytes))
    for sheet_name in list(workbook.sheetnames):
        if sheet_name not in {"event", "JSON字段解析"}:
            del workbook[sheet_name]
    output = BytesIO()
    workbook.save(output)

    parsed = parse_tracking_excel(
        output.getvalue(),
        TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
        physical_schema=_event_physical_schema(),
        datasource_type="mysql",
    )

    assert parsed.editor.default_event_table == "event"
    assert parsed.editor.default_event_name_field == "event"
    assert parsed.editor.default_subject_field == "uid"
    assert parsed.editor.default_event_time_field == "time"
    assert any(field.table_name == "event" and field.field_name == "ext.duration" for field in parsed.editor.fields)


def test_export_infers_event_values_when_default_event_table_is_missing() -> None:
    """
    是什么：验证旧默认事件表未保存时，导出仍能按当前数据源 event 表推断默认字段。
    谁调用：跑测试时 pytest 会找到并执行它。
    做了什么：只保留字段角色和事件值配置，不填 default_event_table，再检查 event sheet 和表映射。
    """
    base = _tracking_config()
    config = base.model_copy(
        update={
            "default_event_table": None,
            "default_subject_field": None,
            "default_event_name_field": None,
            "default_event_time_field": None,
            "tables": [],
            "fields": [
                TenantTrackingFieldDTO(
                    tenant_id=2001,
                    table_name="event",
                    field_name="event",
                    field_comment="业务事件名",
                    field_role="event_name",
                    semantic_type="text",
                ),
                TenantTrackingFieldDTO(
                    tenant_id=2001,
                    table_name="event",
                    field_name="uid",
                    field_comment="用户唯一标识",
                    field_role="subject_id",
                    semantic_type="identifier",
                ),
            ],
        }
    )

    workbook_bytes = tracking_config_excel(config, physical_schema=_event_physical_schema()).getvalue()
    event_rows = _sheet_rows(workbook_bytes, "event")
    table_rows = _sheet_rows(workbook_bytes, "_表映射")
    info_rows = _sheet_rows(workbook_bytes, "_说明")

    assert _headers(workbook_bytes, "event") == ["属性名（必填）", "属性显示名", "属性类型（必填）", "更新方式", "属性说明", "属性标签"]
    assert any(row[0] == "event" and row[1] in (None, "") and row[2] == "文本" for row in event_rows)
    assert not any(row[0] == "event_value" for row in event_rows)
    assert any(row[1] == "event" and row[4] == "uid" and row[5] == "event" for row in table_rows)
    assert any(row[0] == "当前默认" and "事件名字段=event" in str(row[1]) for row in info_rows)


def test_template_contains_datasource_aligned_examples() -> None:
    """
    是什么：验证模板不是空壳，事件表 sheet 会按属性维护格式列出当前数据源字段。
    谁调用：跑测试时 pytest 会找到并执行它。
    做了什么：按真实 event 表生成模板，并检查样例使用当前数据源里的 event/ext 字段。
    """
    workbook_bytes = tracking_config_excel(
        TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
        physical_schema=_event_physical_schema(),
        template_only=True,
    ).getvalue()

    event_rows = _sheet_rows(workbook_bytes, "event")

    assert _headers(workbook_bytes, "event") == ["属性名（必填）", "属性显示名", "属性类型（必填）", "更新方式", "属性说明", "属性标签"]
    assert any(row[0] == "event" and row[2] == "文本" for row in event_rows)
    assert any(row[0] == "ext" and row[2] == "对象组" for row in event_rows)
    assert _headers(workbook_bytes, "JSON字段解析") == JSON_FIELD_PARSING_HEADERS
    assert _sheet_rows(workbook_bytes, "JSON字段解析") == []


def test_template_exports_json_dictionary_fields_only_in_json_overview() -> None:
    """
    验证下载模板时不会丢失当前工作空间已经配置的 JSON 字典字段。
    """
    physical_schema = {
        "event": PhysicalTableInfo(
            table_name="event",
            fields=[
                PhysicalFieldInfo("event", "varchar", "事件名"),
                PhysicalFieldInfo("adinfo", "json", "广告信息"),
                PhysicalFieldInfo("allianceinfo", "json", "联盟信息"),
            ],
        ),
        "user": PhysicalTableInfo(
            table_name="user",
            fields=[
                PhysicalFieldInfo("uid", "varchar", "用户ID"),
                PhysicalFieldInfo("personal", "json", "个人属性"),
            ],
        ),
    }
    config = TenantTrackingConfigDTO(
        tenant_id=2001,
        fields=[
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="event",
                field_name="adinfo.adsetId",
                field_comment="广告组ID",
                semantic_type="text",
                source_field="adinfo",
                json_path="$.adsetId",
            ),
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="event",
                field_name="allianceinfo.allianceid",
                field_comment="联盟ID",
                semantic_type="text",
                source_field="allianceinfo",
                json_path="$.allianceid",
            ),
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="user",
                field_name="personal.nickname",
                field_comment="昵称",
                semantic_type="text",
                source_field="personal",
                json_path="$.nickname",
            ),
        ],
    )

    workbook_bytes = tracking_config_excel(
        config,
        physical_schema=physical_schema,
        template_only=True,
    ).getvalue()

    assert _headers(workbook_bytes, "JSON字段解析") == JSON_FIELD_PARSING_HEADERS
    json_rows = _sheet_rows(workbook_bytes, "JSON字段解析")
    assert ("adinfo", "$.adsetId", "adinfo.adsetId", "文本", "广告组ID") in json_rows
    assert ("allianceinfo", "$.allianceid", "allianceinfo.allianceid", "文本", "联盟ID") in json_rows
    assert ("personal", "$.nickname", "personal.nickname", "文本", "昵称") in json_rows

    event_rows = _sheet_rows(workbook_bytes, "event")
    field_names = {row[0] for row in event_rows}
    assert "adinfo.adsetId" not in field_names
    assert "allianceinfo.allianceid" not in field_names

    user_rows = _sheet_rows(workbook_bytes, "user")
    assert "personal.nickname" not in {row[0] for row in user_rows}

    parsed = parse_tracking_excel(
        workbook_bytes,
        TenantTrackingConfigDTO(
            tenant_id=2001,
            enabled=True,
            default_event_table="event",
        ),
        physical_schema=physical_schema,
        datasource_type="postgresql",
    )
    imported_fields = {
        (field.table_name, field.field_name, field.source_field, field.json_path)
        for field in parsed.editor.fields
    }
    assert ("event", "adinfo.adsetId", "adinfo", "$.adsetId") in imported_fields
    assert ("user", "personal.nickname", "personal", "$.nickname") in imported_fields


def test_json_field_parsing_export_preserves_observed_null_type() -> None:
    field = TenantTrackingFieldDTO(
        tenant_id=2001,
        table_name="event",
        field_name="allianceinfo.power",
        semantic_type="text",
        source_field="allianceinfo",
        json_path="$.power",
        extra_properties={"json_observed_type": "空值"},
    )
    output = tracking_config_excel(
        TenantTrackingConfigDTO(
            tenant_id=2001,
            default_event_table="event",
            fields=[field],
        ),
        physical_schema={
            "event": PhysicalTableInfo(
                "event",
                fields=[PhysicalFieldInfo("allianceinfo", "text")],
            ),
        },
    ).getvalue()

    assert _sheet_rows(output, "JSON字段解析") == [
        ("allianceinfo", "$.power", "allianceinfo.power", "空值", None),
    ]


def test_json_field_parsing_export_rejects_cross_table_source_conflict() -> None:
    config = TenantTrackingConfigDTO(
        tenant_id=2001,
        default_event_table="event",
        fields=[
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="event",
                field_name="payload.event_key",
                semantic_type="text",
                source_field="payload",
                json_path="$.event_key",
            ),
            TenantTrackingFieldDTO(
                tenant_id=2001,
                table_name="user",
                field_name="payload.user_key",
                semantic_type="text",
                source_field="payload",
                json_path="$.user_key",
            ),
        ],
    )
    schema = {
        "event": PhysicalTableInfo(
            "event", fields=[PhysicalFieldInfo("payload", "text")]
        ),
        "user": PhysicalTableInfo(
            "user", fields=[PhysicalFieldInfo("payload", "text")]
        ),
    }

    with pytest.raises(ValueError, match=r"JSON字段解析.*payload.*event.*user"):
        tracking_config_excel(config, physical_schema=schema)


def test_import_excel_replaces_previous_tracking_config() -> None:
    """
    是什么：验证 Excel 是维护源，导入不会把 Excel 已删除的旧事件/字段悄悄合并回来。
    谁调用：跑测试时 pytest 会找到并执行它。
    做了什么：从导出中删掉 pay_success，再用带旧事件的 existing 导入，确认旧事件消失。
    """
    config = _tracking_config()
    workbook_bytes = tracking_config_excel(config, physical_schema=_physical_schema()).getvalue()
    workbook = load_workbook(BytesIO(workbook_bytes))
    sheet = workbook["event_log"]
    for row_index in range(sheet.max_row, 1, -1):
        if sheet.cell(row_index, 11).value in {"pay_success", "refund_success"}:
            sheet.delete_rows(row_index, 1)
    event_parameter_sheet = workbook["事件参数对照"]
    for row_index in range(event_parameter_sheet.max_row, 1, -1):
        if event_parameter_sheet.cell(row_index, 1).value in {"pay_success", "refund_success"}:
            event_parameter_sheet.delete_rows(row_index, 1)
    output = BytesIO()
    workbook.save(output)

    existing = config.model_copy(
        update={
            "event_name_mappings": [
                {"event_name": "legacy_event", "event_display_name": "旧事件"},
                *config.event_name_mappings,
            ],
        }
    )
    parsed = parse_tracking_excel(
        output.getvalue(),
        existing,
        physical_schema=_physical_schema(),
        datasource_type="postgresql",
    )
    event_names = {
        event.get("event_name")
        for event in parsed.editor.event_name_mappings
        if isinstance(event, dict)
    }

    assert "login" in event_names
    assert "pay_success" not in event_names
    assert "legacy_event" not in event_names
    assert not any(field.field_name == "legacy_field" for field in parsed.editor.fields)


def test_import_excel_deleting_field_removes_previous_dictionary_field() -> None:
    """
    是什么：验证 Excel 删除字段行后重新导入，不会把旧字段从 existing 合并回来。
    """
    config = _tracking_config()
    workbook_bytes = tracking_config_excel(config, physical_schema=_physical_schema()).getvalue()
    workbook = load_workbook(BytesIO(workbook_bytes))
    sheet = workbook["JSON字段解析"]
    for row_index in range(sheet.max_row, 1, -1):
        if sheet.cell(row_index, 3).value == "event_props.amount":
            sheet.delete_rows(row_index, 1)
    output = BytesIO()
    workbook.save(output)

    existing = config.model_copy(
        update={
            "fields": [
                *config.fields,
                TenantTrackingFieldDTO(
                    tenant_id=2001,
                    table_name="event_log",
                    field_name="legacy_field",
                    field_comment="旧字段",
                    semantic_type="text",
                ),
            ]
        }
    )
    parsed = parse_tracking_excel(
        output.getvalue(),
        existing,
        physical_schema=_physical_schema(),
        datasource_type="postgresql",
    )
    field_names = {field.field_name for field in parsed.editor.fields}

    assert "event_props.amount" not in field_names
    assert "legacy_field" not in field_names


def test_event_legacy_compact_rows_are_not_imported() -> None:
    """
    是什么：event 表旧紧凑行不再兼容导入。
    """
    parsed = parse_tracking_excel(
        _legacy_business_workbook(
            "event",
            [
                ["physical_field", "event", "事件名", "varchar", "event_name", "text"],
                ["event_value", "event", "", "", "event_name", "text", "", "", "", "", "login", "登录"],
                ["", "", "", "", "", "", "", "", "", "", "pay_success", "支付成功"],
            ],
        ),
        TenantTrackingConfigDTO(tenant_id=2001, enabled=True),
        physical_schema=_event_physical_schema(),
        datasource_type="mysql",
    )

    assert any("event sheet 使用旧 row_type 格式" in warning for warning in parsed.warnings)
    assert not any(event.get("event_name") == "pay_success" for event in parsed.editor.event_name_mappings if isinstance(event, dict))


def test_tracking_prompt_filters_schema_drifted_fields() -> None:
    """
    是什么：物理 schema 漂移后，AI prompt 不再暴露已经失效的字典字段。
    """
    config = _tracking_config()
    config.fields.append(
        TenantTrackingFieldDTO(
            tenant_id=2001,
            table_name="event_log",
            field_name="missing_payload.amount",
            field_comment="失效金额",
            semantic_type="number",
            source_field="missing_payload",
            json_path="$.amount",
            expression="missing expression",
        )
    )

    filtered, validation = filter_tracking_config_for_physical_schema(config, _physical_schema())
    context, summary = build_tracking_prompt_context(filtered, validation.warnings, datasource_type="postgresql")

    assert "event_log.event_props.amount" in context
    assert 'expression=NULLIF(("event_log"."event_props"::jsonb ->> \'amount\'), \'\')::numeric' in context
    assert "`event_log.missing_payload.amount`" not in context
    assert "expression=missing expression" not in context
    assert any("missing_payload" in warning for warning in validation.warnings)
    assert any("schema校验" in item for item in summary)


def test_chart_analysis_and_predict_prompts_share_tracking_semantics() -> None:
    """
    是什么：验证图表生成、分析和预测 Agent 都能看到同一份项目数据字典语义。
    谁调用：跑测试时 pytest 会找到并执行它。
    做了什么：构造问题对象，检查各类 prompt 都包含 tracking_config 与 Data Skill。
    """
    question = AiModelQuestion(
        question="看付费金额趋势",
        sql="select 1",
        fields="[]",
        data="[]",
        tracking_config="<Workspace-Tracking-Rules>支付金额字段 event_props.amount</Workspace-Tracking-Rules>",
        data_skill="Data Skill: 付费金额按事件 pay_success 统计。",
    )

    assert "event_props.amount" in question.chart_user_question()
    assert "pay_success" in question.chart_user_question()
    assert "event_props.amount" in question.analysis_sys_question()
    assert "pay_success" in question.analysis_sys_question()
    assert "event_props.amount" in question.predict_sys_question()
    assert "pay_success" in question.predict_sys_question()


def test_event_name_field_is_text_in_tracking_prompt_context() -> None:
    """
    是什么：验证旧配置里事件名字段若误标为 category，agent 语义出口仍按文本字段表达。
    谁调用：跑测试时 pytest 会找到并执行它。
    做了什么：构造旧库形态的 event_name 字段，检查 prompt 不再输出 type=category。
    """
    config = _tracking_config()
    for field in config.fields:
        if field.field_role == "event_name":
            field.semantic_type = "category"

    context, summary = build_tracking_prompt_context(config)

    assert "`event_log.event_name`; role=event_name; type=text" in context
    assert "role=event_name; type=category" not in context
    assert any("`event_log.event_name`; role=event_name; type=text" in item for item in summary)

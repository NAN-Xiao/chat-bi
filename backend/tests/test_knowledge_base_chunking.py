"""Deterministic standardization and full-replacement chunking tests."""

from __future__ import annotations

import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from apps.knowledge_base.chunking import chunk_knowledge, parse_and_normalize_version
from apps.knowledge_base.normalizers import standardized_content
from apps.knowledge_base.schemas import (
    BusinessKnowledgePayload,
    DocumentPayload,
    EventKnowledgePayload,
    EventParameter,
    JsonFieldKnowledgePayload,
)


def test_same_document_has_stable_normalized_content_and_chunks(tmp_path: Path):
    source = tmp_path / "knowledge.md"
    source.write_text("# 收入\n\n订单收入说明\n\n## SQL\n\n```sql\nselect sum(amount) from orders\n```\n", encoding="utf-8")
    first = parse_and_normalize_version(source)
    second = parse_and_normalize_version(source)
    first_chunks = chunk_knowledge(source=source, chunk_size=40, overlap=8)
    second_chunks = chunk_knowledge(source=source, chunk_size=40, overlap=8)
    assert first.normalized_content == second.normalized_content
    assert [(item.section_path, item.content_hash) for item in first_chunks] == [
        (item.section_path, item.content_hash) for item in second_chunks
    ]


def test_reupload_does_not_merge_missing_old_section():
    chunks = chunk_knowledge(
        DocumentPayload(knowledge_type="DOCUMENT", markdown="# Kept\nnew"),
        chunk_size=1200,
        overlap=150,
    )
    assert all("Removed section" not in chunk.content for chunk in chunks)


def test_all_structured_payloads_have_stable_business_text():
    payloads = [
        BusinessKnowledgePayload(
            knowledge_type="BUSINESS",
            term="收入",
            definition="订单金额总和",
            examples=[{"name": "收入", "question": "收入是多少", "sql": "select sum(amount) from orders"}],
        ),
        EventKnowledgePayload(
            knowledge_type="EVENT",
            event_name="order_paid",
            table_name="events",
            event_name_field="event_name",
            parameters=[EventParameter(name="amount", data_type="decimal")],
        ),
        JsonFieldKnowledgePayload(
            knowledge_type="JSON_FIELD",
            table_name="orders",
            source_field="properties",
            json_path="$.channel",
            field_name="channel",
            data_type="string",
            expression="json_extract(properties, '$.channel')",
        ),
    ]
    for payload in payloads:
        chunks = chunk_knowledge(payload)
        assert chunks
        assert chunks[0].content_hash
        assert chunks[0].token_count > 0


def test_chunk_overlap_and_validation_are_bounded():
    payload = DocumentPayload(knowledge_type="DOCUMENT", markdown="# A\n" + ("word " * 80))
    chunks = chunk_knowledge(payload, chunk_size=50, overlap=10)
    assert len(chunks) > 1
    assert all(len(chunk.content) <= 50 for chunk in chunks)
    assert [chunk.chunk_index for chunk in chunks] == list(range(len(chunks)))


def test_fenced_sql_comments_are_not_treated_as_headings():
    payload = DocumentPayload(
        knowledge_type="DOCUMENT",
        markdown="# 收入\n\n```sql\nselect 1\n# keep this SQL comment\n```\n\n正文",
    )
    parsed = parse_and_normalize_version(payload=payload)
    assert sum("# keep this SQL comment" in content for _, content in parsed.sections) == 1
    assert all(path != "keep this SQL comment" for path, _ in parsed.sections)


def test_structured_content_hash_inputs_are_visible_in_standard_text():
    payload = JsonFieldKnowledgePayload(
        knowledge_type="JSON_FIELD",
        table_name="orders",
        source_field="properties",
        json_path="$.channel",
        field_name="channel",
        display_name="渠道",
        data_type="string",
        expression="json_extract(properties, '$.channel')",
        value_mappings={"1": "广告"},
    )
    content = standardized_content(payload)
    assert "显示名称: 渠道" in content
    assert "值映射" in content


def test_docx_heading_and_hidden_run_are_normalized(tmp_path: Path):
    source = tmp_path / "knowledge.docx"
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:body><w:p><w:pPr><w:pStyle w:val="Heading1"/></w:pPr>'
        '<w:r><w:t>标题</w:t></w:r></w:p>'
        '<w:p><w:r><w:t>公开</w:t></w:r><w:r><w:rPr><w:vanish/></w:rPr>'
        '<w:t>隐藏</w:t></w:r></w:p></w:body></w:document>'
    ).encode()
    with zipfile.ZipFile(source, "w") as archive:
        archive.writestr("word/document.xml", document)
    parsed = parse_and_normalize_version(source)
    assert "# 标题" in parsed.normalized_content
    assert "公开" in parsed.normalized_content
    assert "隐藏" not in parsed.normalized_content


def test_xlsx_worksheets_become_document_sections_and_chunks(tmp_path: Path):
    source = tmp_path / "knowledge.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "指标定义"
    first.append(["指标", "口径"])
    first.append(["收入", "订单实付金额"])
    second = workbook.create_sheet("空表")
    second.append([None])
    third = workbook.create_sheet("SQL 示例")
    third.append(["问题", "SQL"])
    third.append(["收入是多少", "select sum(amount) from orders"])
    workbook.save(source)

    parsed = parse_and_normalize_version(source)
    chunks = chunk_knowledge(source=source, chunk_size=1200, overlap=150)

    assert parsed.source_format == "xlsx"
    assert [path for path, _ in parsed.sections] == ["指标定义", "SQL 示例"]
    assert "收入 | 订单实付金额" in parsed.normalized_content
    assert "select sum(amount) from orders" in parsed.normalized_content
    assert [chunk.section_path for chunk in chunks] == ["指标定义", "SQL 示例"]


def test_invalid_xlsx_is_rejected_with_clear_error(tmp_path: Path):
    source = tmp_path / "broken.xlsx"
    source.write_bytes(b"not-an-excel-workbook")

    with pytest.raises(ValueError, match="Excel 文档格式无效或已损坏"):
        parse_and_normalize_version(source)

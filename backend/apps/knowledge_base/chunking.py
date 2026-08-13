"""Deterministic parsing, standardization, and heading-aware chunking."""

from __future__ import annotations

import hashlib
import re
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from apps.knowledge_base.normalizers import normalize_markdown, standardized_content
from apps.knowledge_base.schemas import DocumentPayload, KnowledgePayload
from common.core.config import settings

_HEADING = re.compile(r"^(#{1,6})\s+(.+?)\s*$")


@dataclass(frozen=True)
class ParsedKnowledge:
    normalized_content: str
    source_format: str
    sections: tuple[tuple[str, str], ...]


@dataclass(frozen=True)
class KnowledgeChunkDraft:
    chunk_index: int
    section_path: str
    content: str
    content_hash: str
    token_count: int
    source_block_id: str | None = None


def parse_and_normalize_version(
    source: str | Path | None = None,
    *,
    payload: KnowledgePayload | None = None,
    file_ext: str | None = None,
    scope: str = "",
) -> ParsedKnowledge:
    """Parse a source file or payload without retaining any previous version text."""
    if payload is not None:
        content = standardized_content(payload, scope=scope)
        source_format = "payload"
    else:
        if source is None:
            raise ValueError("知识源不能为空。")
        path = Path(source)
        extension = (file_ext or path.suffix).lower()
        if extension in {".md", ".markdown"}:
            content = _decode_markdown(path)
            source_format = "markdown"
        elif extension == ".docx":
            content = _decode_docx(path)
            source_format = "docx"
        elif extension == ".xlsx":
            content = _decode_xlsx(path)
            source_format = "xlsx"
        else:
            raise ValueError(f"不支持的知识源格式: {extension}")
        content = normalize_markdown(content)
    sections = tuple(_split_sections(content))
    return ParsedKnowledge(
        normalized_content=content,
        source_format=source_format,
        sections=sections,
    )


def chunk_knowledge(
    payload: KnowledgePayload | None = None,
    *,
    source: str | Path | None = None,
    file_ext: str | None = None,
    scope: str = "",
    chunk_size: int | None = None,
    overlap: int | None = None,
) -> list[KnowledgeChunkDraft]:
    """Produce stable chunks, preserving section paths and bounded overlap."""
    chunk_size = int(settings.KNOWLEDGE_CHUNK_SIZE if chunk_size is None else chunk_size)
    overlap = int(settings.KNOWLEDGE_CHUNK_OVERLAP if overlap is None else overlap)
    if chunk_size <= 0:
        raise ValueError("切片长度必须大于 0。")
    if overlap < 0 or overlap >= chunk_size:
        raise ValueError("切片重叠长度必须小于切片长度。")
    if isinstance(payload, DocumentPayload):
        result: list[KnowledgeChunkDraft] = []
        for block in payload.blocks:
            if not block.enabled:
                continue
            section_text = normalize_markdown(f"# {block.title or '正文'}\n\n{block.markdown}")
            for content in _bounded_chunks(section_text, chunk_size=chunk_size, overlap=overlap):
                result.append(KnowledgeChunkDraft(
                    chunk_index=len(result),
                    section_path=block.title or "正文",
                    content=content,
                    content_hash=hashlib.sha256(content.encode("utf-8")).hexdigest(),
                    token_count=_estimate_tokens(content),
                    source_block_id=block.id,
                ))
        return result
    parsed = parse_and_normalize_version(
        source,
        payload=payload,
        file_ext=file_ext,
        scope=scope,
    )
    result: list[KnowledgeChunkDraft] = []
    for section_path, section_text in parsed.sections:
        for content in _bounded_chunks(section_text, chunk_size=chunk_size, overlap=overlap):
            result.append(
                KnowledgeChunkDraft(
                    chunk_index=len(result),
                    section_path=section_path,
                    content=content,
                    content_hash=hashlib.sha256(content.encode("utf-8")).hexdigest(),
                    token_count=_estimate_tokens(content),
                    source_block_id=None,
                )
            )
    return result


def _decode_markdown(path: Path) -> str:
    data = _read_limited_file(path)
    for encoding in ("utf-8-sig", "utf-8"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _decode_docx(path: Path) -> str:
    _ensure_file_size(path)
    with zipfile.ZipFile(path) as archive:
        info = archive.getinfo("word/document.xml")
        max_bytes = int(settings.KNOWLEDGE_FILE_MAX_BYTES)
        if info.file_size > max_bytes:
            raise ValueError("文档解压后的内容超过允许大小。")
        if info.compress_size <= 0 or info.file_size / info.compress_size > int(settings.KNOWLEDGE_DOCX_MAX_COMPRESSION_RATIO):
            raise ValueError("文档压缩比例异常，已拒绝解析。")
        if sum(item.file_size for item in archive.infolist()) > max_bytes:
            raise ValueError("文档解压后的内容超过允许大小。")
        xml = archive.read(info)
    root = ET.fromstring(xml)
    paragraphs: list[str] = []
    for paragraph in root.iter():
        if _local_name(paragraph.tag) != "p":
            continue
        style = next(
            (
                str(node.attrib.get("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val") or "")
                for node in paragraph.iter()
                if _local_name(node.tag) == "pStyle"
            ),
            "",
        )
        parts: list[str] = []
        for run in paragraph.iter():
            if _local_name(run.tag) != "r":
                continue
            if any(_local_name(node.tag) == "vanish" for node in run.iter()):
                continue
            for node in run.iter():
                name = _local_name(node.tag)
                if name == "t":
                    parts.append(node.text or "")
                elif name == "tab":
                    parts.append("\t")
                elif name in {"br", "cr"}:
                    parts.append("\n")
        value = "".join(parts).strip()
        if value:
            heading_match = re.search(r"heading([1-6])", style, re.IGNORECASE)
            if heading_match:
                value = "#" * int(heading_match.group(1)) + " " + value
            paragraphs.append(value)
    return "\n".join(paragraphs)


def _decode_xlsx(path: Path) -> str:
    """Convert each worksheet into heading-aware Markdown-like text."""
    _ensure_file_size(path)
    _ensure_office_archive_size(path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except (InvalidFileException, KeyError, OSError, zipfile.BadZipFile, ET.ParseError) as exc:
        raise ValueError("Excel 文档格式无效或已损坏。") from exc
    sections: list[str] = []
    try:
        for worksheet in workbook.worksheets:
            rows: list[list[str]] = []
            for row in worksheet.iter_rows(values_only=True):
                values = [_xlsx_cell_text(value) for value in row]
                while values and not values[-1]:
                    values.pop()
                if any(values):
                    rows.append(values)
            if not rows:
                continue
            width = max(len(row) for row in rows)
            normalized = [row + [""] * (width - len(row)) for row in rows]
            title = re.sub(r"[\r\n]+", " ", worksheet.title).strip() or "工作表"
            sections.append(f"# {title}\n\n" + "\n".join(
                " | ".join(row) for row in normalized
            ))
    finally:
        workbook.close()
    if not sections:
        raise ValueError("Excel 文档没有可用内容。")
    return "\n\n".join(sections)


def _xlsx_cell_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[\r\n]+", " ", str(value)).strip()


def _ensure_office_archive_size(path: Path) -> None:
    max_bytes = int(settings.KNOWLEDGE_FILE_MAX_BYTES)
    max_ratio = int(settings.KNOWLEDGE_DOCX_MAX_COMPRESSION_RATIO)
    try:
        with zipfile.ZipFile(path) as archive:
            items = archive.infolist()
            if sum(item.file_size for item in items) > max_bytes:
                raise ValueError("文档解压后的内容超过允许大小。")
            if any(
                item.file_size > 0
                and (item.compress_size <= 0 or item.file_size / item.compress_size > max_ratio)
                for item in items
            ):
                raise ValueError("文档压缩比例异常，已拒绝解析。")
    except zipfile.BadZipFile as exc:
        raise ValueError("Excel 文档格式无效或已损坏。") from exc


def _ensure_file_size(path: Path) -> None:
    if path.stat().st_size > int(settings.KNOWLEDGE_FILE_MAX_BYTES):
        raise ValueError("知识源文件超过允许大小。")


def _read_limited_file(path: Path) -> bytes:
    _ensure_file_size(path)
    return path.read_bytes()


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _split_sections(content: str) -> Iterable[tuple[str, str]]:
    lines = content.splitlines()
    heading_stack: list[str] = []
    section_lines: list[str] = []
    section_path = "正文"
    fenced = False

    def flush() -> tuple[str, str] | None:
        text = normalize_markdown("\n".join(section_lines))
        body = normalize_markdown("\n".join(section_lines[1:])) if section_lines and _HEADING.match(section_lines[0]) else text
        if not text or (section_lines and _HEADING.match(section_lines[0]) and not body):
            return None
        return section_path, text

    for line in lines:
        match = None if fenced else _HEADING.match(line)
        if match:
            previous = flush()
            if previous:
                yield previous
            level = len(match.group(1))
            heading_stack[:] = heading_stack[: level - 1]
            heading_stack.append(match.group(2))
            section_path = " / ".join(heading_stack)
            section_lines = [line]
        else:
            section_lines.append(line)
        if re.match(r"^\s*(```|~~~)", line):
            fenced = not fenced
    previous = flush()
    if previous:
        yield previous


def _bounded_chunks(text: str, *, chunk_size: int, overlap: int) -> Iterable[str]:
    if len(text) <= chunk_size:
        yield text
        return
    start = 0
    length = len(text)
    while start < length:
        end = min(length, start + chunk_size)
        protected = _protected_range_at(text, end)
        if protected is not None and protected[1] > end:
            end = protected[1]
        if end < length:
            boundary = max(text.rfind("\n", start, end), text.rfind(" ", start, end))
            if boundary > start + chunk_size // 2:
                end = boundary
        chunk = text[start:end].strip()
        if chunk:
            yield chunk
        if end >= length:
            return
        next_start = max(start + 1, end - overlap)
        protected_start = _protected_range_at(text, next_start)
        if protected_start is not None:
            next_start = (
                protected_start[1]
                if protected_start[0] <= start
                else protected_start[0]
            )
        if next_start <= start:
            next_start = end
        start = next_start


def _protected_range_at(text: str, position: int) -> tuple[int, int] | None:
    """Return the full fenced-code span containing a character position."""
    fence = re.compile(r"^\s*(```|~~~)", re.MULTILINE)
    matches = list(fence.finditer(text))
    if len(matches) < 2:
        return None
    for index in range(0, len(matches) - 1, 2):
        start = matches[index].start()
        end = matches[index + 1].end()
        if start <= position < end:
            return start, end
    return None


def _estimate_tokens(text: str) -> int:
    return max(1, (len(text) + 3) // 4)
